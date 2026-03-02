import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


def _venv_python() -> str:
    return sys.executable


def _excel_script() -> str:
    pkg_root = Path(__file__).resolve().parents[1]
    return str(pkg_root / "scripts" / "generate_excel.py")


class TestGenerateExcel(unittest.TestCase):
    def test_excel_headers_and_mapping(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            input_path = td_path / "testcases.json"
            output_path = td_path / "out.xlsx"

            testcases = [
                {
                    "id": "需求A_202602280001",
                    "title": "登录_凭据验证_正确凭据登录成功",
                    "feature": "登录",
                    "type": "正向",
                    "preconditions": "1、用户已注册",
                    "steps": "1、打开登录页\n2、输入账号密码\n3、点击登录",
                    "expected_result": "1、登录成功并跳转首页",
                    "priority": "P1",
                }
            ]
            input_path.write_text(json.dumps(testcases, ensure_ascii=False), encoding="utf-8")

            subprocess.check_call(
                [
                    _venv_python(),
                    _excel_script(),
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                ],
                timeout=30,
            )

            wb = load_workbook(output_path, read_only=True, data_only=True)
            ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active

            headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
            self.assertEqual(
                headers,
                [
                    "编号",
                    "用例标题",
                    "级别",
                    "预置条件",
                    "操作步骤",
                    "测试预期内容",
                    "执行结果",
                    "执行人",
                    "执行日期",
                    "备注",
                ],
            )

            self.assertEqual(ws.cell(row=2, column=1).value, "需求A_202602280001")
            self.assertEqual(ws.cell(row=2, column=2).value, "登录_凭据验证_正确凭据登录成功")
            self.assertEqual(ws.cell(row=2, column=3).value, "P1")
            self.assertIn(ws.cell(row=2, column=7).value, (None, ""))

            wb.close()

    def test_excel_with_quoted_text(self):
        """字段值包含中文引号「」和转义双引号时正常生成。"""
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            input_path = td_path / "testcases.json"
            output_path = td_path / "out.xlsx"

            testcases = [
                {
                    "id": "需求B_202602280001",
                    "title": "任务_按钮文案_显示去完成",
                    "feature": "任务",
                    "type": "正向",
                    "preconditions": "1、用户已登录",
                    "steps": "1、进入任务列表\n2、点击「去完成」按钮",
                    "expected_result": '1、按钮文案为「去完成」\n2、跳转至任务详情页',
                    "priority": "P1",
                }
            ]
            input_path.write_text(json.dumps(testcases, ensure_ascii=False), encoding="utf-8")

            subprocess.check_call(
                [_venv_python(), _excel_script(), "--input", str(input_path), "--output", str(output_path)],
                timeout=30,
            )

            wb = load_workbook(output_path, read_only=True, data_only=True)
            ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active
            self.assertIn("去完成", ws.cell(row=2, column=5).value)
            self.assertIn("去完成", ws.cell(row=2, column=6).value)
            wb.close()

    def test_excel_invalid_json_exits_with_message(self):
        """testcases.json 格式非法时脚本应返回非零退出码并输出提示。"""
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            input_path = td_path / "testcases.json"
            output_path = td_path / "out.xlsx"

            input_path.write_text('[{"expected_result": "按钮文案为"去完成""}]', encoding="utf-8")

            result = subprocess.run(
                [_venv_python(), _excel_script(), "--input", str(input_path), "--output", str(output_path)],
                capture_output=True, text=True, timeout=30,
            )
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("JSON 格式无效", result.stderr)


if __name__ == "__main__":
    unittest.main()
