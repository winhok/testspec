#!/usr/bin/env python3
"""
TestSpec Excel 用例生成脚本：根据 testcases.json 生成 .xlsx 测试用例文档。

用法：
    python generate_excel.py --input testcases.json --output artifacts/cases.xlsx
"""
import argparse
import json
import logging
import os
import sys

logger = logging.getLogger(__name__)

# 表头定义：(名称, 列宽)
COLUMNS = [
    ("编号", 18),
    ("用例标题", 45),
    ("级别", 8),
    ("预置条件", 25),
    ("操作步骤", 40),
    ("测试预期内容", 35),
    ("执行结果", 12),
    ("执行人", 10),
    ("执行日期", 14),
    ("备注", 20),
]


def create_excel_with_openpyxl(test_cases: list, output_path: str) -> None:
    """使用 openpyxl 创建 Excel 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("需要安装 openpyxl，请运行：pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    try:
        sheet = wb.active
        sheet.title = "测试用例"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, (header, _) in enumerate(COLUMNS, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (_, width) in enumerate(COLUMNS, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, tc in enumerate(test_cases, 2):
            tc_id = tc.get("id") or tc.get("case_id") or f"TC-{row_idx - 1:03d}"
            title = tc.get("title") or tc.get("name", "")
            priority = tc.get("priority", "P1")
            preconditions = tc.get("preconditions", "")
            steps = tc.get("steps", "")
            expected = tc.get("expected_result", tc.get("expected", ""))

            sheet.cell(row=row_idx, column=1, value=tc_id)
            sheet.cell(row=row_idx, column=2, value=title)
            sheet.cell(row=row_idx, column=3, value=priority)
            sheet.cell(row=row_idx, column=4, value=preconditions)
            sheet.cell(row=row_idx, column=5, value=steps)
            sheet.cell(row=row_idx, column=6, value=expected)
            sheet.cell(row=row_idx, column=7, value="")
            sheet.cell(row=row_idx, column=8, value="")
            sheet.cell(row=row_idx, column=9, value="")
            sheet.cell(row=row_idx, column=10, value="")

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Excel test cases from JSON")
    parser.add_argument("--input", "-i", required=True, help="Path to testcases.json")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    try:
        with open(args.input, encoding="utf-8-sig") as f:
            test_cases = json.load(f)
    except FileNotFoundError:
        logger.error("文件不存在: %s", args.input)
        sys.exit(1)
    except json.JSONDecodeError as e:
        logger.error("%s JSON 格式无效（行 %s 列 %s）。", args.input, e.lineno, e.colno)
        logger.error("常见原因：字符串值中包含未转义的双引号。请检查并将 \" 转义为 \\\" 或使用「」替代。")
        sys.exit(1)

    if not isinstance(test_cases, list):
        logger.error("JSON 应为用例数组")
        sys.exit(1)

    create_excel_with_openpyxl(test_cases, args.output)
    print(f"已生成：{args.output}")


if __name__ == "__main__":
    main()
