import json
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


def _venv_python() -> str:
    repo_root = Path(__file__).resolve().parents[2]
    return str(repo_root / ".venv" / "bin" / "python")


def _excel_script() -> str:
    pkg_root = Path(__file__).resolve().parents[1]
    return str(pkg_root / "scripts" / "generate_excel.py")


class TestGenerateExcel(unittest.TestCase):
    def test_excel_headers_and_mapping(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            input_path = td_path / "testcases.json"
            output_path = td_path / "out.xlsx"

            testcases = [
                {
                    "id": "需求A_202602280001",
                    "title": "登录_凭据验证_正确凭据登录成功",
                    "feature": "登录",
                    "type": "正向",
                    "preconditions": "1、用户已注册",
                    "steps": "1、打开登录页\n2、输入账号密码\n3、点击登录",
                    "expected_result": "1、登录成功并跳转首页",
                    "priority": "P0",
                }
            ]
            input_path.write_text(json.dumps(testcases, ensure_ascii=False), encoding="utf-8")

            subprocess.check_call(
                [
                    _venv_python(),
                    _excel_script(),
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                ]
            )

            wb = load_workbook(output_path, read_only=True, data_only=True)
            ws = wb["测试用例"] if "测试用例" in wb.sheetnames else wb.active

            headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
            self.assertEqual(
                headers,
                [
                    "编号",
                    "用例标题",
                    "级别",
                    "预置条件",
                    "操作步骤",
                    "测试预期内容",
                    "执行结果",
                    "执行人",
                    "执行日期",
                    "备注",
                ],
            )

            self.assertEqual(ws.cell(row=2, column=1).value, "需求A_202602280001")
            self.assertEqual(ws.cell(row=2, column=2).value, "登录_凭据验证_正确凭据登录成功")
            self.assertEqual(ws.cell(row=2, column=3).value, "P0")
            self.assertIn(ws.cell(row=2, column=7).value, (None, ""))

            wb.close()


if __name__ == "__main__":
    unittest.main()
