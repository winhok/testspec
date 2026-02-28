#!/usr/bin/env python3
"""
TestSpec Excel 用例生成脚本：根据 testcases.json 生成 .xlsx 测试用例文档。

用法：
    python generate_excel.py --input testcases.json --output artifacts/cases.xlsx
"""
import argparse
import json
import os
import sys


def create_excel_with_openpyxl(test_cases: list, output_path: str) -> None:
    """使用 openpyxl 创建 Excel 文件。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误：需要安装 openpyxl，请运行：pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "测试用例"

    headers = [
        "编号",
        "用例标题",
        "级别",
        "预置条件",
        "操作步骤",
        "测试预期内容",
        "执行结果",
        "执行人",
        "执行日期",
        "备注",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [18, 45, 8, 25, 40, 35, 12, 10, 14, 20]
    for col_idx, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, tc in enumerate(test_cases, 2):
        tc_id = tc.get("id") or tc.get("case_id") or f"TC-{row_idx - 1:03d}"
        title = tc.get("title") or tc.get("name", "")
        priority = tc.get("priority", "P1")
        preconditions = tc.get("preconditions", "")
        steps = tc.get("steps", "")
        expected = tc.get("expected_result", tc.get("expected", ""))

        sheet.cell(row=row_idx, column=1, value=tc_id)
        sheet.cell(row=row_idx, column=2, value=title)
        sheet.cell(row=row_idx, column=3, value=priority)
        sheet.cell(row=row_idx, column=4, value=preconditions)
        sheet.cell(row=row_idx, column=5, value=steps)
        sheet.cell(row=row_idx, column=6, value=expected)
        sheet.cell(row=row_idx, column=7, value="")
        sheet.cell(row=row_idx, column=8, value="")
        sheet.cell(row=row_idx, column=9, value="")
        sheet.cell(row=row_idx, column=10, value="")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate Excel test cases from JSON")
    parser.add_argument("--input", "-i", required=True, help="Path to testcases.json")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.input, encoding="utf-8-sig") as f:
        test_cases = json.load(f)

    if not isinstance(test_cases, list):
        print("错误：JSON 应为用例数组", file=sys.stderr)
        sys.exit(1)

    create_excel_with_openpyxl(test_cases, args.output)
    print(f"已生成：{args.output}")


if __name__ == "__main__":
    main()
